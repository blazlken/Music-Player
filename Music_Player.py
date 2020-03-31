import tkinter as tk
from pygame import mixer
import multiprocessing
import os
import openpyxl

class MainApplication(tk.Frame):
    def __init__(self, parent, *args, **kwargs):
        tk.Frame.__init__(self, parent, *args, **kwargs)
        self.parent = parent

        self.parent.title("Music Player")
        self.parent.geometry("250x300")

        self.CurrentPlaylist = tk.StringVar(value="Current Playlist: " + Playlist.wb.sheetnames[Playlist.playNum])
        Current_Playlist = tk.Label(self.parent, textvariable=self.CurrentPlaylist)
        Current_Playlist.place(relx=0.5, rely=0.1, anchor='n')

        Now_Playing = tk.Label(self.parent, text="Now Playing:")
        Now_Playing.place(relx=0.5, rely=0.30, anchor='s')

        self.CurrentTrack = tk.StringVar(value=str(Playlist.wb[Playlist.wb.sheetnames[Playlist.playNum]]['B2'].value)[:-4])
        self.CurTrackName = tk.Label(self.parent, textvariable=self.CurrentTrack)
        self.CurTrackName.place(relx=0.5, rely=0.37, anchor='s')

        Next_Up = tk.Label(self.parent, text="Next Track:")
        Next_Up.place(relx=0.5, rely=0.85, anchor='center')

        self.NextTrack = tk.StringVar(value=str(Playlist.wb[Playlist.wb.sheetnames[Playlist.playNum]]['B3'].value)[:-4])
        self.NextTrackName = tk.Label(self.parent, textvariable=self.NextTrack)
        self.NextTrackName.place(relx=0.5, rely=0.90, anchor='n')

        Next_Track_Btn = tk.Button(self.parent, text="Next Track", command=self.Next_Track)
        Next_Track_Btn.place(relx=0.6, rely=0.55, anchor='w')

        Prev_Track_Btn = tk.Button(self.parent, text="Previous Track", command=self.Prev_Track)
        Prev_Track_Btn.place(relx=0.4, rely=0.55, anchor='e')

        Next_P_Btn = tk.Button(self.parent, text="Next Playlist", command=self.Next_Playlist)
        Next_P_Btn.place(relx=0.6, rely=0.65, anchor='w')

        Prev_P_Btn = tk.Button(self.parent, text="Previous Playlist", command=self.Prev_Playlist)
        Prev_P_Btn.place(relx=0.4, rely=0.65, anchor='e')

        self.Looping = tk.IntVar()
        Repeat = tk.Checkbutton(self.parent, text='Loop', variable=self.Looping, onvalue=-1)
        Repeat.place(relx=0.5, rely=0.7625, anchor='center')

        self.PlayBtnImg = tk.PhotoImage(file=r'Images\\play.png').subsample(3, 3)
        play_or_resume = tk.Button(self.parent, image=self.PlayBtnImg, command=self.Play_Resume)
        play_or_resume.place(relx=0.5, rely=0.55, anchor='center')

        self.PauseBtnImg = tk.PhotoImage(file=r'Images\\pause.png').subsample(3, 3)
        pause = tk.Button(self.parent, image=self.PauseBtnImg, command=self.Pause)
        pause.place(relx=0.5, rely=0.65, anchor='center')
        pass

    def Next_Playlist(self):
        Playlist.trackNum = 2

        if Playlist.playNum >= (Playlist.WSNum - 1):
            Playlist.playNum = 0
            pass
        else:
            Playlist.playNum += 1

        self.ActivePlaylist = Playlist.wb.sheetnames[Playlist.playNum]

        Playlist.FirstEmptyFunc()

        Playlist.LoadTrack()

        self.TrackNameUpdate()

        mixer.music.play(self.Looping.get())
        pass

    def Prev_Playlist(self):
        Playlist.trackNum = 2

        if Playlist.playNum >= (Playlist.WSNum - 1):
            Playlist.playNum = 0
            pass
        else:
            Playlist.playNum += 1

        self.ActivePlaylist = Playlist.wb.sheetnames[Playlist.playNum]

        Playlist.FirstEmptyFunc()

        Playlist.LoadTrack()

        self.TrackNameUpdate()

        mixer.music.play(self.Looping.get())
        pass

    def Next_Track(self):
        Playlist.trackNum += 1
        if ('B' + str(Playlist.trackNum)) == Playlist.FirstEmpty:
            Playlist.trackNum = 2
            pass
        else:
            pass

        Playlist.LoadTrack()

        self.TrackNameUpdate()

        mixer.music.play(self.Looping.get())
        pass

    def Prev_Track(self):
        Playlist.trackNum -= 1
        if ('B' + str(Playlist.trackNum)) == 'B1':
            Playlist.trackNum = int(Playlist.FirstEmpty[-1:]) - 1
            pass
        else:
            pass

        Playlist.LoadTrack()

        self.TrackNameUpdate()

        mixer.music.play(self.Looping.get())
        pass
    
    def TrackNameUpdate(self):
        self.CurrentTrack.set(str(Playlist.wb[Playlist.wb.sheetnames[Playlist.playNum]]['B' + str(Playlist.trackNum)].value)[:-4])
        self.CurrentPlaylist.set('Current Playlist: ' + Playlist.wb.sheetnames[Playlist.playNum])

        if (Playlist.trackNum + 1) == int(Playlist.FirstEmpty[-1:]):
            self.NextTrack.set(str(Playlist.wb[Playlist.wb.sheetnames[Playlist.playNum]]['B2'].value)[:-4])
            pass
        else:
            self.NextTrack.set(str(Playlist.wb[Playlist.wb.sheetnames[Playlist.playNum]]['B' + str(Playlist.trackNum + 1)].value)[:-4])
            pass

        pass

    def Play_Resume(self):
        mixer.music.unpause()

        if mixer.music.get_busy() == False:
            mixer.music.play(self.Looping.get())
            pass
        
        pass

    def Pause(self):
        mixer.music.pause()
        pass

         

class Playlist():
    def __init__(self):
        self.filepath = os.path.dirname(os.path.abspath(__file__)) + "\\Music\\Playlists.xlsx"
        self.playNum = 0
        self.trackNum = 2

        self.wb = openpyxl.load_workbook(self.filepath,False)
        
        self.WSNum = len(self.wb.sheetnames)

        self.FirstEmptyFunc()
        self.LoadTrack()

        pass

    def FirstEmptyFunc(self):
        self.FirstEmpty = None
        num = 2
        while self.FirstEmpty is None:
            if self.wb[self.wb.sheetnames[self.playNum]]['B' + str(num)].value is None:
                self.FirstEmpty = 'B' + str(num)
                pass
            else:
                num += 1
                pass
        
    def LoadTrack(self):
        mixer.music.stop()
        mixer.music.load('Music\\Tracks\\' + str(self.wb[self.wb.sheetnames[self.playNum]]['B' + str(self.trackNum)].value))
        pass


if __name__ == "__main__":
    mixer.init()
    Playlist = Playlist()

    root = tk.Tk()
    MainApp = MainApplication(root)
    MainApp.pack(side="top", fill="both", expand=False)
    root.mainloop()